from pathlib import Path
from typing import Hashable

import numpy as np
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter

from . import utils


def run():
    exam = utils.get_user_input('시험(0: LEET, 1: PSAT 5급, 2: PSAT 7급): ', 0, int)
    num_students = utils.get_user_input('학생 수: ', 1000, int)
    input_excel_file = utils.get_user_input('답안 선택률 엑셀 파일명: ', '선택률', str)
    output_excel_file = utils.get_user_input('가상 답안 엑셀 파일명: ', '가상_답안', str)

    if exam == 0:
        fake_answer = LeetFakeAnswer(num_students, input_excel_file, output_excel_file)
    elif exam == 1:
        fake_answer = PsatHaengsiFakeAnswer(num_students, input_excel_file, output_excel_file)
    else:
        fake_answer = PsatChilgeubFakeAnswer(num_students, input_excel_file, output_excel_file)

    output_excel = fake_answer.output_excel
    sheet_data = fake_answer.get_sheet_data()
    answer_count_sheet_data = fake_answer.get_answer_count_set_data()

    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        for sheet_name, df in sheet_data.items():
            df.to_excel(writer, sheet_name=sheet_name)

        for sheet_name, df in answer_count_sheet_data.items():
            df.to_excel(writer, sheet_name=f'answer_count_{sheet_name}')

    fake_answer.adjust_column_widths()
    print(f"✅ '{output_excel}' 파일에 모든 시트를 저장했습니다.")


class LeetFakeAnswer:
    BASE_DIR = Path('D:/projects/#generate_virtual_answers')

    SUBJECTS = ['subject_0', 'subject_1']
    SUBJECTS_PLUS_TOTAL = SUBJECTS + ['total']
    BASE_MEANS = {'subject_0': 45, 'subject_1': 60}
    BASE_STDS = {'subject_0': 10, 'subject_1': 10}
    PROBLEM_COUNTS = {'subject_0': 30, 'subject_1': 40}

    RANK_LABEL = 'standard_score'
    METRICS = ['correct_count', 'percentage_score', 'standard_score']
    STAT_KEYS = ['max', 't10', 't25', 't50', 'min', 'avg', 'std']
    BIN_WIDTH = 10
    SCORE_LEVEL_ADJUSTMENTS = {'top': 1.2, 'mid': 1.0, 'low': 0.8}

    def __init__(self, num_students: int, input_excel_file: str, output_excel_file: str):
        self.num_students = num_students
        self.input_excel = self.BASE_DIR / f'{input_excel_file}.xlsx'
        self.output_excel = self.BASE_DIR / f'{output_excel_file}.xlsx'

        # 입력 데이터프레임
        self.input_answer_distribution_df = pd.read_excel(
            self.input_excel, index_col=[2, 3], sheet_name='answer_distribution')
        self.group_distribution_df = pd.read_excel(self.input_excel, index_col=0, sheet_name='group_distribution')
        self.correct_answer_df = self.input_answer_distribution_df[['answer']]
        self.adjusted_answer_rate_df = self.get_adjusted_answer_rate_df()
        self.input_catalog_df = self.get_input_catalog_df()

        # 가상 데이터 생성
        self.ids = [f'fake{i:04}' for i in range(1, num_students + 1)]
        self.student_answer_df = self.get_student_answer_df()
        self.all_answers, self.all_scores, self.all_stats, self.temp_totals = self.get_all_fake_data()

        # 학생 정보
        self.student_df = self.get_student_df()

        # 출력 데이터프레임
        self.catalog_df = self.get_catalog_df()
        self.statistics_df = self.get_statistics_df()
        self.frequency_df = self.get_frequency_df()

    def get_adjusted_answer_rate_df(self) -> pd.DataFrame:
        adjusted_rows = []

        for (subject, number), dist_row in self.input_answer_distribution_df.iterrows():
            counts = np.array([dist_row.get(f'count_{i}', 0) for i in range(1, 6)])
            count_sum = dist_row.get('count_sum', 0)
            if count_sum > 0:
                base_rates = counts / count_sum
            else:
                base_rates = np.ones(5) / 5  # 균등분포 fallback

            correct = self.correct_answer_df.loc[(subject, number), 'answer']

            for level, ratio in self.SCORE_LEVEL_ADJUSTMENTS.items():
                adjusted = base_rates.copy()
                adjusted[correct - 1] *= ratio  # 정답 선택지 확률 조정

                # Normalize
                adjusted = self.safe_normalize(adjusted)  # 안전한 정규화 적용

                adjusted_rows.append({
                    'subject': subject,
                    'number': number,
                    'score_level': level,
                    **{f'rate_{i + 1}': adjusted[i] for i in range(5)}
                })

        adjusted_df = pd.DataFrame(adjusted_rows)
        adjusted_df.set_index(['subject', 'number', 'score_level'], inplace=True)
        return adjusted_df

    @staticmethod
    def safe_normalize(arr, tol=1e-8):
        arr = np.clip(arr, 0, None)  # 음수 제거
        total = arr.sum()

        if total == 0:
            # 균등 분포로 반환
            return np.ones_like(arr) / len(arr)

        arr = arr / total
        correction = 1.0 - arr.sum()

        # correction이 무시해도 되는 수준이 아니면 가장 큰 값에 더해줌
        if abs(correction) > tol:
            max_idx = arr.argmax()
            arr[max_idx] += correction

        return arr

    def get_input_catalog_df(self):
        xls = pd.ExcelFile(self.input_excel)
        if 'catalog' in xls.sheet_names:
            return pd.read_excel(self.input_excel, index_col=0, header=[0, 1], sheet_name='catalog')
        return pd.DataFrame()

    def get_student_answer_df(self):
        if not self.input_catalog_df.empty:
            df = self.input_catalog_df
            level_0 = df.columns.get_level_values(0)
            answer_columns = df.columns[level_0.isin(self.SUBJECTS)]
            df = df[answer_columns]
            return df
        else:
            sampled_combos = self.sample_score_combination()
            all_answers = self.generate_grouped_fake_answers(sampled_combos)
            # all_answers = {}
            # for _, row in sampled_combos.iterrows():
            #     row_answers = []
            #     for subject in self.SUBJECTS:
            #         group = row[subject]
            #         subject_answers = self.generate_answers_for_group(subject, group)
            #         df = self.input_answer_distribution_df
            #         df_subject = df[df.index.get_level_values('subject') == subject]
            #         student_answers = self.generate_student_answers(df_subject)
            #         all_answers[subject] = student_answers

            answer_cols = pd.MultiIndex.from_tuples([
                (subject, number) for subject, count in self.PROBLEM_COUNTS.items() for number in range(1, count + 1)
            ])
            answers_combined = np.hstack([all_answers[subject] for subject in self.SUBJECTS])
            answer_df = pd.DataFrame(answers_combined, index=self.ids, columns=answer_cols)
            answer_df.index.name = 'serial'
            return answer_df

    def sample_score_combination(self, seed: int = None) -> pd.DataFrame:
        if seed is not None:
            np.random.seed(seed)

        combo_probs = self.group_distribution_df.value_counts(normalize=True).reset_index()
        combo_probs.columns = ['total', 'subject_0', 'subject_1', 'prob']
        samples = combo_probs.sample(n=self.num_students, weights='prob', replace=True)
        return samples[['total', 'subject_0', 'subject_1']].reset_index(drop=True)

    # @staticmethod
    # def softmax(x, temp=1.0):
    #     x = np.array(x)
    #     x /= temp
    #     e_x = np.exp(x - np.max(x))
    #     return e_x / e_x.sum()
    #
    # def adjust_probs(self, probs, level):
    #     if level == 'top':
    #         return self.softmax(probs, temp=0.5)  # 상위권: 정답률 강조
    #     elif level == 'low':
    #         return self.softmax(probs[::-1], temp=0.5)[::-1]  # 하위권: 오답률 강조
    #     else:  # 중위권
    #         return probs / np.sum(probs)

    def generate_grouped_fake_answers(self, score_tiers_df):
        answer_data = {}
        grouped = score_tiers_df.groupby(['total'] + self.SUBJECTS)

        for group_key, group_df in grouped:
            idxs = group_df.index
            n = len(group_df)

            for (subject, number) in self.adjusted_answer_rate_df.index.droplevel('score_level').unique():
                # 해당 과목의 점수대 선택
                subject_tier = group_key[0] if subject == 'total' else group_key[int(subject[-1]) + 1]

                # 정답률 조정된 선택지 비율 가져오기
                try:
                    rate_row = self.adjusted_answer_rate_df.loc[(subject, number, subject_tier)]
                    probs = rate_row.values
                except KeyError:
                    probs = np.ones(5) / 5  # fallback to uniform distribution if missing

                # 정답 생성
                answers = np.random.choice(range(1, 6), size=n, p=probs)
                if (subject, number) not in answer_data:
                    answer_data[(subject, number)] = np.zeros(len(score_tiers_df), dtype=int)
                answer_data[(subject, number)][idxs] = answers

        # 멀티인덱스 컬럼 적용
        answer_df = pd.DataFrame(answer_data)
        answer_df.columns = pd.MultiIndex.from_tuples(answer_df.columns, names=['subject', 'number'])
        return answer_df

    # def generate_student_answers(self, df: pd.DataFrame) -> np.array:
    #     probs_matrix = df[['count_1', 'count_2', 'count_3', 'count_4', 'count_5']].div(
    #         df[['count_1', 'count_2', 'count_3', 'count_4', 'count_5']].sum(axis=1), axis=0)
    #     # 각 문제마다 num_students 명의 선택지를 생성
    #     student_answers = np.array([
    #         np.random.choice([1, 2, 3, 4, 5], size=self.num_students, p=probs_matrix.iloc[q_idx].values)
    #         for q_idx in range(len(df))
    #     ]).T  # 전치하여 (num_students, num_questions) 형태로 만듦
    #
    #     return student_answers

    def get_all_fake_data(self):
        df = self.input_answer_distribution_df
        all_answers, all_scores, all_stats, temp_totals = {}, {}, {}, {}

        for subject in self.SUBJECTS:
            df_subject = df[df.index.get_level_values('subject') == subject]
            correct_answers = df_subject['answer'].to_numpy()
            student_answers = self.student_answer_df[subject].to_numpy()
            correct_counts, percentage_scores, standard_scores = self.compute_scores(
                subject, correct_answers, student_answers)

            all_answers[subject] = student_answers
            all_scores[subject] = {
                'correct_count': correct_counts,
                'percentage_score': percentage_scores,
                'standard_score': standard_scores,
            }
            all_stats[subject] = {k: self.compute_statistics(all_scores[subject][k]) for k in self.METRICS}

        all_stats['total'] = {}
        for metric in self.METRICS:
            total = np.sum([all_scores[s][metric] for s in self.SUBJECTS], axis=0)
            temp_totals[metric] = total
            all_stats['total'][metric] = self.compute_statistics(total)

        return all_answers, all_scores, all_stats, temp_totals

    def compute_scores(
            self, subject: str, correct_answers: np.array, student_answers: np.array
    ) -> tuple[np.array, np.array, np.array]:
        correct_counts = (student_answers == correct_answers).sum(axis=1)
        percentage_scores = np.round(correct_counts / len(correct_answers) * 100, 1)
        base_mean = self.BASE_MEANS.get(subject, 50)
        base_std = self.BASE_STDS.get(subject, 10)
        standard_scores = np.round(
            (correct_counts - correct_counts.mean()) / correct_counts.std() * base_std + base_mean, 1)

        return correct_counts, percentage_scores, standard_scores

    def get_student_df(self):
        passwords = [f'{np.random.randint(0, 10000):04}' for _ in range(self.num_students)]

        df = pd.read_excel(self.input_excel, index_col=0, sheet_name='aspiration_data')
        df = self.clean_choice_dataframe(df)
        choices = self.assign_choices(df)

        student_data = {
            'password': passwords,
            'aspiration_1': choices['aspiration_1'].values,
            'aspiration_2': choices['aspiration_2'].values,
        }
        student_df = pd.DataFrame(student_data, index=self.ids)
        student_df.index.name = 'serial'
        return student_df

    def assign_choices(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        전체 학생 수에 대해 1지망/2지망 배정:
        - 절반은 조합 비율에 따라
        - 절반은 개별 분포(1지망/2지망 각각)에 따라 샘플링
        """

        # 조합 비율 계산 및 조합 기반 샘플링
        n_combo = self.num_students // 2
        combo_probs = df.groupby(['aspiration_1', 'aspiration_2']).size() / len(df)
        combo_samples = combo_probs.sample(n=n_combo, weights=combo_probs.values, replace=True)
        combo_choices = combo_samples.reset_index()[['aspiration_1', 'aspiration_2']]

        # 개별 분포 계산 및 개별 분포 기반 샘플링
        n_indiv = self.num_students - n_combo
        indiv_choices_dict = {}
        for i in range(1, 3):
            aspiration_df = df[f'aspiration_{i}']
            aspiration_df_filtered = aspiration_df[aspiration_df != '미선택']
            probs = aspiration_df_filtered.value_counts(normalize=True)
            indiv_choices_dict[f'aspiration_{i}'] = np.random.choice(probs.index, size=n_indiv, p=probs.values)
        indiv_choices = pd.DataFrame(indiv_choices_dict)

        choices = pd.concat([combo_choices, indiv_choices], ignore_index=True)
        return choices

    def get_catalog_df(self):
        if not self.input_catalog_df.empty:
            return self.input_catalog_df
        else:
            catalog = {}
            for metric in self.METRICS:
                for subject in self.SUBJECTS_PLUS_TOTAL:
                    if subject == 'total':
                        catalog[(metric, 'total')] = self.temp_totals[metric]
                    else:
                        catalog[(metric, subject)] = self.all_scores[subject][metric]

            # 등수 계산 및 catalog에 추가
            for subject in self.SUBJECTS_PLUS_TOTAL:
                if subject == 'total':
                    values = self.temp_totals[self.RANK_LABEL]
                else:
                    values = self.all_scores[subject][self.RANK_LABEL]
                catalog[('rank', subject)] = self.get_descending_ranks(values)

            score_df = pd.DataFrame(catalog, index=self.student_df.index)

            # 열 위치 조정: rank / correct_count / percentage_score / standard_score
            level_0 = score_df.columns.get_level_values(0)
            rank_cols = score_df.columns[level_0 == 'rank']
            cc_cols = score_df.columns[level_0 == 'correct_count']
            other_cols = score_df.columns[(~level_0.isin(['rank', 'correct_count']))]
            new_order = list(rank_cols) + list(cc_cols) + list(other_cols)
            score_df = score_df[new_order]

            for subject in reversed(self.SUBJECTS_PLUS_TOTAL):
                rank_subject_label = (self.RANK_LABEL, subject)
                quantiles = score_df[rank_subject_label].quantile([0.27, 0.73])

                def classify_group(score):
                    if score <= quantiles[0.27]:
                        return 'low'
                    elif score <= quantiles[0.73]:
                        return "mid"
                    else:
                        return 'top'

                group_label = ('group', subject)
                group_col = score_df[rank_subject_label].apply(classify_group)
                score_df.insert(0, group_label, group_col)

            student_df = pd.DataFrame(self.student_df)
            student_df.columns = pd.MultiIndex.from_tuples([
                ('student', 'password'), ('student', 'aspiration_1'), ('student', 'aspiration_2')
            ])

            catalog_df = student_df.join(score_df, how='left').join(self.student_answer_df, how='left')
            return catalog_df

    def get_sheet_data(self):
        return {
            'correct_answer': self.correct_answer_df,
            'catalog': self.catalog_df,
            'aspiration': self.get_aspiration_df(),
            'statistics': self.statistics_df,
            'frequency': self.frequency_df,
        }

    def get_aspiration_df(self) -> pd.DataFrame:
        aspiration_1_counts = self.student_df['aspiration_1'].value_counts().sort_index()
        aspiration_2_counts = self.student_df['aspiration_2'].value_counts().sort_index()
        aspiration_total_counts = aspiration_1_counts.add(aspiration_2_counts, fill_value=0)

        aspiration_df = pd.DataFrame({
            'aspiration_1': aspiration_1_counts,
            'aspiration_2': aspiration_2_counts,
            'total': aspiration_total_counts,
        }).fillna(0).astype(int)
        aspiration_df.index.name = 'university'

        aspiration_df = aspiration_df.reset_index()
        aspiration_df['sort_key'] = aspiration_df['university'].apply(lambda x: 'ㅎㅎㅎ' if x == '기타 대학' else x)
        aspiration_df = aspiration_df.sort_values('sort_key').drop(columns='sort_key')
        aspiration_df = aspiration_df.set_index('university')

        return aspiration_df

    def get_statistics_df(self) -> pd.DataFrame:
        stat_rows = [
            ((m, s), [round(self.all_stats[s][m].get(k, 0), 1) for k in self.STAT_KEYS])
            for m in self.METRICS for s in self.SUBJECTS_PLUS_TOTAL
        ]
        stat_index = pd.MultiIndex.from_tuples([row[0] for row in stat_rows], names=['metric', 'subject'])
        return pd.DataFrame([row[1] for row in stat_rows], index=stat_index, columns=self.STAT_KEYS)

    def get_frequency_df(self) -> pd.DataFrame:
        freq_rows = []
        for metric in self.METRICS:
            for subject in self.SUBJECTS_PLUS_TOTAL:
                if subject == 'total':
                    values = self.temp_totals[metric]
                else:
                    values = self.all_scores[subject][metric]

                bin_width = 1 if metric == 'correct_count' else self.BIN_WIDTH
                freq_df = self.generate_frequency_table(values, subject, bin_width)

                for _, row in freq_df.iterrows():
                    freq_rows.append((metric, row['subject'], row['bin'], row['frequency']))

        freq_df = pd.DataFrame(freq_rows, columns=['metric', 'subject', 'bin', 'freq'])
        freq_df.set_index(['metric', 'subject', 'bin'], inplace=True)

        return freq_df

    def get_answer_count_set_data(self) -> dict[Hashable, pd.DataFrame]:
        all_group = self.catalog_df.copy()
        all_group[('group', 'total')] = 'all'
        student_groups = pd.concat([all_group, self.catalog_df], axis=0).groupby(('group', 'total'))

        sheet_data = {}
        for rank_type, group_df in student_groups:
            choice_counts = []
            for subject in self.SUBJECTS:
                response_array = np.array(group_df.xs(subject, axis=1, level=0))
                answer_distribution = self.calculate_answer_distribution(response_array, subject)
                choice_counts.append(answer_distribution.reset_index(drop=True))
            sheet_data[rank_type] = self.get_answer_count_df(choice_counts)

        return sheet_data

    def get_answer_count_df(self, choice_counts):
        answer_count_df = pd.concat(choice_counts, ignore_index=True)

        subject = answer_count_df.pop('subject')
        number = answer_count_df.pop('number')
        answer_count_df.index = pd.MultiIndex.from_arrays([subject, number])

        answer_count_df = self.correct_answer_df.join(answer_count_df)

        answer_count_df['answer_rate'] = answer_count_df.apply(
            lambda row: row[f'rate_{int(row["answer"])}'], axis=1
        )
        cols = list(answer_count_df.columns)
        cols.insert(cols.index('answer') + 1, cols.pop(cols.index('answer_rate')))
        answer_count_df = answer_count_df[cols]

        return answer_count_df

    def adjust_column_widths(self):
        wb = load_workbook(self.output_excel)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col_idx, col in enumerate(ws.iter_cols(), 1):
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max_length + 2
        wb.save(self.output_excel)

    @staticmethod
    def clean_choice_dataframe(df: pd.DataFrame, fill_value='미선택') -> pd.DataFrame:
        df = df.rename(columns=lambda x: x.strip().lower().replace(' ', '_'))
        df = df[['aspiration_1', 'aspiration_2']].copy()
        df = df.fillna(fill_value)
        df['aspiration_1'] = df['aspiration_1'].replace('', fill_value)
        df['aspiration_2'] = df['aspiration_2'].replace('', fill_value)
        return df

    @staticmethod
    def compute_statistics(values: np.array) -> dict[str, int | float]:
        return {
            'max': values.max(),
            't10': np.percentile(values, 90),
            't25': np.percentile(values, 75),
            't50': np.percentile(values, 50),
            'min': values.min(),
            'avg': values.mean(),
            'std': values.std(),
        }

    @staticmethod
    def generate_frequency_table(values: np.ndarray, subject: str, bin_width: int) -> pd.DataFrame:
        """구간 기반 도수분포표 생성 함수

        Args:
            values (np.ndarray): 학생 점수 데이터
            subject (str): 과목 필드 (혹은 'total')
            bin_width (int, optional): 구간 너비. 기본값은 10

        Returns:
            pd.DataFrame: [과목, 구간, 빈도수] 형태의 도수분포표
        """
        series = pd.Series(values)
        min_val, max_val = int(series.min()), int(series.max()) + 1
        start_bin = (min_val // bin_width) * bin_width
        end_bin = ((max_val // bin_width) + 1) * bin_width
        bins = list(range(start_bin, end_bin + 1, bin_width))

        # 구간 라벨 생성
        if bin_width == 1:
            labels = [b for b in bins[:-1]]
        else:
            labels = [f'{b}점 이상~{b + bin_width}점 미만' for b in bins[:-1]]

        binned = pd.cut(series, bins=bins, labels=labels, right=False)
        freq = binned.value_counts().sort_index()
        freq_df = freq.rename_axis('bin').reset_index(name='frequency')
        freq_df.insert(0, 'subject', subject)

        return freq_df

    @staticmethod
    def calculate_answer_distribution(response_array: np.ndarray, subject: str) -> pd.DataFrame:
        """문항별 학생 선택률(①~⑤)을 계산하여 DataFrame으로 반환

        Args:
            response_array (np.ndarray): (학생 수, 문항 수) 형태의 배열
            subject: 과목 코드

        Returns:
            pd.DataFrame: 각 문항별 선택률 (1~5번 보기 기준)
        """
        num_questions = response_array.shape[1]
        result = []

        for idx in range(num_questions):
            choices = response_array[:, idx]
            counts = pd.Series(choices).value_counts().reindex([1, 2, 3, 4, 5], fill_value=0)
            total = counts.sum()
            rates = counts / total
            row = counts.values.tolist() + [total] + rates.round(4).values.tolist()
            result.append(row)

        answer_count_df = pd.DataFrame(
            result,
            columns=[
                'count_1', 'count_2', 'count_3', 'count_4', 'count_5', 'count_sum',
                'rate_1', 'rate_2', 'rate_3', 'rate_4', 'rate_5',
            ],
            index=[i + 1 for i in range(num_questions)],
        )
        answer_count_df.insert(0, 'subject', subject)
        answer_count_df.insert(1, 'number', answer_count_df.index)

        return answer_count_df

    @staticmethod
    def get_descending_ranks(values: np.ndarray) -> np.ndarray:
        # 동점자에게 동일한 등수 부여
        sorted_idx = np.argsort(-values)
        sorted_values = values[sorted_idx]
        ranks = np.empty_like(values, dtype=int)

        current_rank = 1
        for i in range(len(values)):
            if i > 0 and sorted_values[i] != sorted_values[i - 1]:
                current_rank = i + 1
            ranks[sorted_idx[i]] = current_rank
        return ranks


class PsatHaengsiFakeAnswer(LeetFakeAnswer):
    SUBJECTS = ['subject_0', 'subject_1', 'subject_2', 'subject_3']
    BASE_MEANS = {'subject_0': 50, 'subject_1': 50, 'subject_2': 50, 'subject_3': 50}
    BASE_STDS = {'subject_0': 10, 'subject_1': 10, 'subject_2': 10, 'subject_3': 10}
    RANK_LABEL = 'percentile_score'

    def get_temp_totals_and_update_all_stats(self) -> dict[str, np.array]:
        self.all_stats['total'] = {}
        temp_totals = {m: np.sum([self.all_scores[s][m] for s in self.SUBJECTS[1:]], axis=0) for m in self.METRICS}
        for metric in self.METRICS:
            self.all_stats['total'][metric] = self.compute_statistics(temp_totals[metric])
        return temp_totals

    def get_student_df(self):
        passwords = [f'{np.random.randint(0, 10000):04}' for _ in range(self.num_students)]

        df = pd.read_excel(self.input_excel, sheet_name='category')
        total_applicants = df['applicants'].sum()
        df['rate'] = df['applicants'] / total_applicants
        categories = df[['unit'], ['department']].tolist()
        probabilities = df['rate'].tolist()
        category_data = np.random.choice(categories, size=self.num_students, p=probabilities)

        student_data = {
            'password': passwords,
            'unit': '',
            'department': '',
        }
        return pd.DataFrame(student_data, index=self.ids)

    def get_sheet_data(self):
        return {
            'correct_answer': self.correct_answer_df,
            'catalog': self.catalog_df,
            'category': self.get_category_df(),
            'statistics': self.statistics_df,
            'frequency': self.frequency_df,
        }

    def get_category_df(self):
        pass


class PsatChilgeubFakeAnswer(PsatHaengsiFakeAnswer):
    SUBJECTS = ['subject_1', 'subject_2', 'subject_3']

    def get_temp_totals_and_update_all_stats(self) -> dict[str, np.array]:
        return super(PsatHaengsiFakeAnswer).get_temp_totals_and_update_all_stats()
