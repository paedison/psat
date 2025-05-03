from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter

from . import utils

BASE_DIR = Path('D:/projects/#generate_virtual_answers')
DEFAULT_SUBJECTS = {
    0: ['subject_0', 'subject_1'],  # LEET
    1: ['subject_0', 'subject_1', 'subject_2', 'subject_3'],  # PSAT 5급
    2: ['subject_1', 'subject_2', 'subject_3'],  # PSAT 7급
}

STANDARD_BASE_MEANS = {
    0: {'subject_0': 45, 'subject_1': 60},
    1: {'subject_0': 50, 'subject_1': 50, 'subject_2': 50, 'subject_3': 50},
    2: {'subject_1': 50, 'subject_2': 50, 'subject_3': 50},
}

STANDARD_BASE_STDS = {
    0: {'subject_0': 10, 'subject_1': 10},
    1: {'subject_0': 10, 'subject_1': 10, 'subject_2': 10, 'subject_3': 10},
    2: {'subject_1': 10, 'subject_2': 10, 'subject_3': 10},
}

METRICS = ['correct_count', 'percentage_score', 'standard_score']
STAT_KEYS = ['max', 't10', 't25', 't50', 'min', 'avg', 'std']
BIN_WIDTH = 10


def run():
    exam = utils.get_user_input('시험(0: LEET, 1: PSAT 5급, 2: PSAT 7급): ', 0, int)
    num_students = utils.get_user_input('학생 수: ', 1000, int)
    input_excel_file = utils.get_user_input('답안 선택률 엑셀 파일명: ', '선택률', str)
    output_excel_file = utils.get_user_input('가상 답안 엑셀 파일명: ', '가상_답안', str)

    subjects = DEFAULT_SUBJECTS.get(exam)
    subjects_plus_total = subjects + ['total']
    input_excel = BASE_DIR / f'{input_excel_file}.xlsx'
    output_excel = BASE_DIR / f'{output_excel_file}.xlsx'

    # 결과 저장용
    all_correct_answers, all_answers, all_scores, all_stats = get_virtual_data(
        input_excel, exam, subjects, num_students)

    # 학생 정보
    ids = [f'dummy{i:04}' for i in range(1, num_students + 1)]
    passwords = [f'{np.random.randint(0, 10000):04}' for _ in range(num_students)]
    aspiration_1, aspiration_2 = get_aspiration_data(input_excel, num_students)
    student_df = pd.DataFrame(
        {'password': passwords, 'aspiration_1': aspiration_1, 'aspiration_2': aspiration_2}, index=ids)

    # 전체 통계 및 점수 합산
    all_stats['total'] = {}
    temp_totals = {m: np.sum([all_scores[s][m] for s in subjects], axis=0) for m in METRICS}
    for metric in METRICS:
        all_stats['total'][metric] = compute_statistics(temp_totals[metric])

    # 정답 / 답안 / 지망 선택 현황 / 통계 / 전체 답안 선택률 / 도수분포표 데이터프레임
    correct_answer_df = get_correct_answer_df(all_correct_answers)
    answer_df = get_answer_df(subjects, all_answers, ids)
    aspiration_df = get_aspiration_df(student_df)
    stat_df = get_stat_df(subjects_plus_total, all_stats)
    freq_df = get_freq_df(subjects_plus_total, all_scores, temp_totals)
    answer_count_df = get_answer_count_df(subjects, all_answers)

    # 성적 데이터프레임
    label_for_rank = 'correct_count' if exam else 'standard_score'
    score_df = get_score_df(student_df, subjects, subjects_plus_total, all_scores, temp_totals)
    for subject in reversed(subjects_plus_total):
        # 성적/답안 데이터프레임에 그룹 컬럼 추가
        column_label = ('group', subject)
        group_col = get_group_column(score_df, label_for_rank, subject, column_label)
        score_df.insert(3, column_label, group_col)
        answer_df.insert(0, column_label, group_col)

    # 성적대별 답안 선택률 데이터프레임 세트
    group_choice_counts = get_group_choice_counts(answer_df, subjects)

    # 시트 정보
    sheet_dict = {
        'correct_answer': correct_answer_df,
        'score': score_df,
        'answer': answer_df,
        'aspiration': aspiration_df,
        'statistics': stat_df,
        'frequency': freq_df,
        'answer_count': answer_count_df,
    }

    # Excel 저장
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        for sheet_name, df in sheet_dict.items():
            df.to_excel(writer, sheet_name=sheet_name)

        # 상위권, 중위권, 하위권 시트 저장
        for group_name in ['top', 'mid', 'low']:
            combined_group_df = pd.concat(group_choice_counts[group_name], ignore_index=True)
            change_index_to_subject_number(combined_group_df)
            combined_group_df.to_excel(writer, sheet_name=f'answer_count_{group_name}')

    wb = load_workbook(output_excel)
    adjust_column_widths(wb)
    wb.save(output_excel)

    print(f"✅ '{output_excel}' 파일에 모든 시트를 저장했습니다.")


def get_virtual_data(input_excel, exam, subjects, num_students):
    all_correct_answers, all_answers, all_scores, all_stats = [], {}, {}, {}

    for subject in subjects:
        df = pd.read_excel(input_excel, sheet_name=subject)
        correct_answers = df['정답'].tolist()
        response_array = generate_student_answers(df, num_students)
        correct_count_array, percentage_score_array, standard_score_array = compute_scores(
            correct_answers, response_array, exam, subject)

        all_answers[subject] = response_array
        all_scores[subject] = {
            'correct_count': correct_count_array,
            'percentage_score': percentage_score_array,
            'standard_score': standard_score_array,
        }

        all_correct_answers.extend([(subject, idx + 1, ans) for idx, ans in enumerate(correct_answers)])
        all_stats[subject] = {k: compute_statistics(all_scores[subject][k]) for k in METRICS}

    return all_correct_answers, all_answers, all_scores, all_stats


def generate_student_answers(df, num_students):
    probs_matrix = df[['①', '②', '③', '④', '⑤']].div(df[['①', '②', '③', '④', '⑤']].sum(axis=1), axis=0)
    # 각 문제마다 num_students 명의 선택지를 생성
    response_array = np.array([
        np.random.choice([1, 2, 3, 4, 5], size=num_students, p=probs_matrix.iloc[q_idx].values)
        for q_idx in range(len(df))
    ]).T  # 전치하여 (num_students, num_questions) 형태로 만듦

    return response_array


def compute_scores(correct_answers, response_array, exam, subject):
    correct_count_array = (response_array == np.array(correct_answers)).sum(axis=1)
    percentage_score_array = np.round(correct_count_array / len(correct_answers) * 100, 1)
    base_mean = STANDARD_BASE_MEANS[exam].get(subject, 50)
    base_std = STANDARD_BASE_STDS[exam].get(subject, 10)
    standard_score_array = np.round(
        (correct_count_array - correct_count_array.mean()) / correct_count_array.std() * base_std + base_mean, 1)

    return correct_count_array, percentage_score_array, standard_score_array


def compute_statistics(values):
    return {
        'max': values.max(),
        't10': np.percentile(values, 90),
        't25': np.percentile(values, 75),
        't50': np.percentile(values, 50),
        'min': values.min(),
        'avg': values.mean(),
        'std': values.std(),
    }


def get_aspiration_data(input_excel, num_students):
    univ_df = pd.read_excel(input_excel, sheet_name='university')

    total_applicants = univ_df['applicants'].sum()
    univ_df['rate'] = univ_df['applicants'] / total_applicants
    universities = univ_df['university'].tolist()
    probabilities = univ_df['rate'].tolist()

    aspiration_1 = np.random.choice(universities, size=num_students, p=probabilities)
    aspiration_2 = []
    for fc in aspiration_1:
        available_univs = [u for u in universities if u != fc]
        available_probs = [univ_df.loc[univ_df['university'] == u, 'rate'].values[0] for u in available_univs]
        total = sum(available_probs)
        norm_probs = [p / total for p in available_probs]
        aspiration_2.append(np.random.choice(available_univs, p=norm_probs))

    return aspiration_1, aspiration_2


def get_correct_answer_df(all_correct_answers):
    correct_answer_df = pd.DataFrame(all_correct_answers, columns=['subject', 'number', 'answer'])
    subject = correct_answer_df.pop('subject')
    number = correct_answer_df.pop('number')
    correct_answer_df.index = pd.MultiIndex.from_arrays([subject, number])
    return correct_answer_df


def get_answer_df(subjects, all_answers, ids):
    answer_cols = pd.MultiIndex.from_tuples([
        (s, i + 1) for s in subjects for i in range(all_answers[s].shape[1])
    ])
    answers_combined = np.hstack([all_answers[s] for s in subjects])
    answer_df = pd.DataFrame(answers_combined, index=ids, columns=answer_cols)
    answer_df.index.name = 'serial'
    return answer_df


def get_aspiration_df(student_df):
    aspiration_1_counts = student_df['aspiration_1'].value_counts().sort_index()
    aspiration_2_counts = student_df['aspiration_2'].value_counts().sort_index()
    aspiration_total_counts = aspiration_1_counts.add(aspiration_2_counts, fill_value=0)

    aspiration_df = pd.DataFrame({
        'aspiration_1': aspiration_1_counts,
        'aspiration_2': aspiration_2_counts,
        'total': aspiration_total_counts,
    }).fillna(0).astype(int)
    aspiration_df.index.name = 'university'

    aspiration_df = aspiration_df.reset_index()
    aspiration_df['sort_key'] = aspiration_df['university'].apply(lambda x: 'ㅎㅎㅎ' if x == '기타 대학' else x)
    aspiration_df = aspiration_df.sort_values('sort_key').drop(columns='sort_key')
    aspiration_df = aspiration_df.set_index('university')

    return aspiration_df


def get_stat_df(subjects_plus_total, all_stats):
    stat_rows = [
        ((m, s), [round(all_stats[s][m].get(k, 0), 1) for k in STAT_KEYS])
        for m in METRICS for s in subjects_plus_total
    ]
    stat_index = pd.MultiIndex.from_tuples([row[0] for row in stat_rows], names=['metric', 'subject'])
    return pd.DataFrame([row[1] for row in stat_rows], index=stat_index, columns=STAT_KEYS)


def get_freq_df(subjects_plus_total, all_scores, temp_totals):
    freq_rows = []
    for metric in METRICS:
        for subject in subjects_plus_total:
            values = all_scores[subject][metric] if subject != 'total' else temp_totals[metric]
            bin_width = 1 if metric == 'correct_count' else BIN_WIDTH
            freq_df = generate_frequency_table(values, subject, bin_width=bin_width)
            for _, row in freq_df.iterrows():
                freq_rows.append((metric, row['subject'], row['bin'], row['frequency']))

    freq_df = pd.DataFrame(freq_rows, columns=['metric', 'subject', 'bin', 'freq'])
    freq_df.set_index(['metric', 'subject', 'bin'], inplace=True)

    return freq_df


def generate_frequency_table(values: np.ndarray, subject: str, bin_width: int = 10) -> pd.DataFrame:
    """구간 기반 도수분포표 생성 함수

    Args:
        values (np.ndarray): 학생 점수 데이터
        subject (str): 과목 필드 (혹은 'total')
        bin_width (int, optional): 구간 너비. 기본값은 10

    Returns:
        pd.DataFrame: [과목, 구간, 빈도수] 형태의 도수분포표
    """
    series = pd.Series(values)
    min_val, max_val = int(series.min()), int(series.max()) + 1
    start_bin = (min_val // bin_width) * bin_width
    end_bin = ((max_val // bin_width) + 1) * bin_width
    bins = list(range(start_bin, end_bin + 1, bin_width))

    # 구간 라벨 생성
    if bin_width == 1:
        labels = [b for b in bins[:-1]]
    else:
        labels = [f'{b}점 이상~{b + bin_width}점 미만' for b in bins[:-1]]

    binned = pd.cut(series, bins=bins, labels=labels, right=False)
    freq = binned.value_counts().sort_index()
    freq_df = freq.rename_axis('bin').reset_index(name='frequency')
    freq_df.insert(0, 'subject', subject)

    return freq_df


def get_answer_count_df(subjects, all_answers):
    choice_counts = []
    for subject in subjects:
        response_array = all_answers[subject]
        answer_distribution = calculate_answer_distribution(response_array, subject)
        choice_counts.append(answer_distribution.reset_index(drop=True))

    answer_count_df = pd.concat(choice_counts, ignore_index=True)
    change_index_to_subject_number(answer_count_df)

    return answer_count_df


def change_index_to_subject_number(df):
    subject = df.pop('subject')
    number = df.pop('number')
    df.index = pd.MultiIndex.from_arrays([subject, number])


def calculate_answer_distribution(response_array: np.ndarray, subject: str) -> pd.DataFrame:
    """문항별 학생 선택률(①~⑤)을 계산하여 DataFrame으로 반환

    Args:
        response_array (np.ndarray): (학생 수, 문항 수) 형태의 배열
        subject: 과목 코드

    Returns:
        pd.DataFrame: 각 문항별 선택률 (1~5번 보기 기준)
    """
    num_questions = response_array.shape[1]
    result = []

    for idx in range(num_questions):
        choices = response_array[:, idx]
        counts = pd.Series(choices).value_counts().reindex([1, 2, 3, 4, 5], fill_value=0)
        total = counts.sum()
        row = counts.round(4).values.tolist() + [total]
        result.append(row)

    answer_count_df = pd.DataFrame(
        result,
        columns=['count_1', 'count_2', 'count_3', 'count_4', 'count_5', 'count_sum'],
        index=[i + 1 for i in range(num_questions)],
    )
    answer_count_df.insert(0, 'subject', subject)
    answer_count_df.insert(1, 'number', answer_count_df.index)

    return answer_count_df


def get_score_df(student_df: pd.DataFrame, subjects, subjects_plus_total, all_scores, temp_totals):
    score_data = {}
    for metric in METRICS:
        for subject in subjects:
            score_data[(metric, subject)] = all_scores[subject][metric]
        score_data[(metric, 'total')] = temp_totals[metric]

        # 등수 계산 및 score_data에 추가
        for subject in subjects_plus_total:
            values = all_scores[subject]['correct_count'] if subject != 'total' else temp_totals['correct_count']
            score_data[('rank', subject)] = get_descending_ranks(values)

    score_df = pd.DataFrame(score_data, index=student_df.index)
    student_df.columns = pd.MultiIndex.from_tuples([
        ('student', 'password'), ('student', 'aspiration_1'), ('student', 'aspiration_2')
    ])
    score_df = student_df.join(score_df, how='left')

    return score_df


def get_descending_ranks(values: np.ndarray) -> np.ndarray:
    # 동점자에게 동일한 등수 부여
    sorted_idx = np.argsort(-values)
    sorted_values = values[sorted_idx]
    ranks = np.empty_like(values, dtype=int)

    current_rank = 1
    for i in range(len(values)):
        if i > 0 and sorted_values[i] != sorted_values[i - 1]:
            current_rank = i + 1
        ranks[sorted_idx[i]] = current_rank
    return ranks


def get_group_column(score_df, quantile_score_label, subject, group_column_label):
    quantile_column_label = (quantile_score_label, subject)
    quantiles = score_df[quantile_column_label].quantile([0.27, 0.73])

    def classify_group(score):
        if score <= quantiles[0.27]:
            return 'low'
        elif score <= quantiles[0.73]:
            return "mid"
        else:
            return 'top'

    score_df[group_column_label] = score_df[quantile_column_label].apply(classify_group)
    return score_df.pop(group_column_label)


def get_group_choice_counts(answer_df, subjects):
    group_choice_counts = {}
    answer_cols = [col for col in answer_df.columns if col[0] != 'group']
    student_groups = answer_df.groupby(('group', 'total'))

    for group_name, group_df in student_groups:
        subject_dict = {}

        for subject in subjects:
            subject_cols = [col for col in answer_cols if col[0] == subject]
            group_subject_df = group_df[subject_cols]

            count_data = []

            for number in sorted({col[1] for col in subject_cols}):
                col = (subject, number)
                value_counts = group_subject_df[col].value_counts().sort_index()
                counts = [value_counts.get(i, 0) for i in range(1, 6)]
                count_sum = sum(counts)
                count_data.append([subject, number] + counts + [count_sum])

            subject_df = pd.DataFrame(count_data, columns=[
                'subject', 'number',
                'count_1', 'count_2', 'count_3', 'count_4', 'count_5',
                'count_sum'
            ])
            subject_dict[subject] = subject_df

        group_choice_counts[group_name] = subject_dict

    return group_choice_counts


def adjust_column_widths(wb):
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col_idx, col in enumerate(ws.iter_cols(), 1):
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max_length + 2
