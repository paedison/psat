import os
import random
import shutil

from django.core.management.base import BaseCommand
from openpyxl import Workbook, load_workbook

from a_psat.models import Problem


class Command(BaseCommand):
    help = "Extract problems by subject and year range."

    def add_arguments(self, parser):
        parser.add_argument('start_year', type=int, help="Start year of the problem extraction")
        parser.add_argument('end_year', type=int, help="End year of the problem extraction")
        parser.add_argument('problem_count', type=int, help="Number of problems to extract")

    def handle(self, *args, **options):
        start_year = options['start_year']
        end_year = options['end_year']
        problem_count = options['problem_count']
        subjects = ['언어', '자료', '상황']

        target_folder = 'a_psat/data/selected_problems'
        if not os.path.exists(target_folder):
            os.makedirs(target_folder)

        worksheet_number_list = [0]
        wb_filepath = f'{target_folder}/problem_list.xlsx'
        if os.path.exists(wb_filepath):
            wb = load_workbook(wb_filepath)
            for sheetname in wb.sheetnames:
                worksheet_number_list.append(int(sheetname))
            new_worksheet_number = max(worksheet_number_list) + 1
            ws = wb.create_sheet(title=str(new_worksheet_number))
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = '1'
            new_worksheet_number = 1

        # 기존 엑셀 파일에서 문제 리스트 읽기
        extracted_problem_ids = self.load_existing_problem_ids(wb)

        all_selected_problems = []
        for subject in subjects:
            problems = Problem.objects.filter(
                subject=subject, year__range=(start_year, end_year)).exclude(
                id__in=extracted_problem_ids).exclude(exam='입시')
            selected_problems = random.sample(list(problems), min(problem_count, len(problems)))

            # 추출된 문제 ID를 중복 방지를 위해 기록
            extracted_problem_ids.update(problem.id for problem in selected_problems)

            all_selected_problems.extend(selected_problems)

        exam_order = {'민경': 1, '칠급': 2, '외시': 3, '행시': 4}
        all_selected_problems_sorted = sorted(
            all_selected_problems, key=lambda prob: (prob.subject, exam_order.get(prob.exam, 5)))

        # 엑셀 파일로 저장
        self.save_to_excel(all_selected_problems_sorted, ws)
        wb.save(wb_filepath)
        self.stdout.write(f"문제가 '{target_folder}/{wb_filepath}' 파일로 저장되었습니다.")

        # 이미지 파일 복사 및 정리
        self.organize_image_files(all_selected_problems_sorted, target_folder, new_worksheet_number)

    @staticmethod
    def load_existing_problem_ids(workbook) -> set:
        """기존 엑셀 파일에서 추출된 문제의 ID를 읽어옵니다."""
        extracted_ids = set()
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(min_row=2, values_only=True):  # 첫 번째 행은 헤더이므로 건너뜁니다.
                problem_id = row[0]  # 문제 ID는 첫 번째 열에 있다고 가정합니다.
                extracted_ids.add(problem_id)
        return extracted_ids

    @staticmethod
    def save_to_excel(problems, worksheet):
        worksheet.append(["ID", "일련번호", "연도", "시험", "과목", "책형", "번호", "정답", "발문"])
        for problem in problems:
            worksheet.append([
                problem.id, problem.reference2,
                problem.year, problem.exam, problem.subject, problem.paper_type,
                problem.number, problem.answer, problem.question,
            ])

    def organize_image_files(self, problems, target_folder, worksheet_number):
        base_image_path = "static/image/PSAT"
        folder = os.path.join(target_folder, str(worksheet_number))
        if not os.path.exists(folder):
            os.makedirs(folder)

        # 과목별로 정렬된 문제에 대해 sorted_number를 증가시키면서 파일 복사
        subject_counters = {}

        for sorted_number, problem in enumerate(problems, start=1):
            subject = problem.subject
            year = problem.year
            exam = problem.exam
            number = problem.number

            # 각 과목별로 번호를 독립적으로 관리
            if subject not in subject_counters:
                subject_counters[subject] = 1
            sorted_number = subject_counters[subject]

            # 원본 파일명
            original_file_1 = f"PSAT{year}{exam}{subject}{number:02}-1.png"
            original_file_2 = f"PSAT{year}{exam}{subject}{number:02}-2.png"

            # 원본 파일 경로
            original_file_path_1 = os.path.join(base_image_path, str(year), original_file_1)
            original_file_path_2 = os.path.join(base_image_path, str(year), original_file_2)

            # 새로운 파일명
            new_file_1 = f"{subject}{sorted_number:02}_{original_file_1}"
            new_file_2 = f"{subject}{sorted_number:02}_{original_file_2}"

            # 새로운 파일 경로
            new_file_path_1 = os.path.join(folder, new_file_1)
            new_file_path_2 = os.path.join(folder, new_file_2)

            # 파일이 존재하면 복사
            if os.path.exists(original_file_path_1):
                shutil.copy(original_file_path_1, new_file_path_1)
                # self.stdout.write(f"Copied {original_file_1} to {new_file_path_1}")

            if os.path.exists(original_file_path_2):
                shutil.copy(original_file_path_2, new_file_path_2)
                # self.stdout.write(f"Copied {original_file_2} to {new_file_path_2}")

            # 과목별로 sorted_number 증가
            subject_counters[subject] += 1
